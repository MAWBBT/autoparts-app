import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
from datetime import datetime
import io

# Колонки таблицы (используются и в on_change редактора)
_TABLE_COLUMNS = ["Наименование", "Артикул", "Бренд", "Цена, руб.", "Кол-во"]
_MAIN_EDITOR_KEY = "main_editor"


def _empty_positions_table():
    return pd.DataFrame(columns=_TABLE_COLUMNS)


def _coerce_editor_cell(column: str, value):
    """Приводит значение из JSON редактора к типу колонки."""
    if value is None:
        return None
    if column == "Кол-во":
        try:
            return int(float(value))
        except (ValueError, TypeError):
            return value
    if column == "Цена, руб.":
        try:
            return float(str(value).replace(",", ".").replace(" ", ""))
        except (ValueError, TypeError):
            return value
    return value


def _on_main_editor_change():
    """
    Синхронизация правок без строки вида
    session_state.main_data = data_editor(session_state.main_data),
    из‑за которой Streamlit пересоздаёт виджет и правка «съедается» до второго ввода.
    См. https://github.com/streamlit/streamlit/issues/7749
    """
    delta = st.session_state.get(_MAIN_EDITOR_KEY)
    if not isinstance(delta, dict):
        return

    df = st.session_state.main_data

    for row_idx, changes in (delta.get("edited_rows") or {}).items():
        ri = int(row_idx)
        if ri < 0 or ri >= len(df):
            continue
        for col, val in changes.items():
            if col in df.columns:
                df.iat[ri, df.columns.get_loc(col)] = _coerce_editor_cell(col, val)

    deleted = delta.get("deleted_rows") or []
    if deleted:
        to_drop = [df.index[int(i)] for i in deleted if 0 <= int(i) < len(df)]
        if to_drop:
            df.drop(index=to_drop, inplace=True)
            df.reset_index(drop=True, inplace=True)

    added = delta.get("added_rows") or []
    if added:
        base = df.reset_index(drop=True)
        rows = []
        for raw in added:
            if not isinstance(raw, dict):
                continue
            rows.append(
                {c: _coerce_editor_cell(c, raw.get(c)) for c in _TABLE_COLUMNS}
            )
        if rows:
            st.session_state.main_data = pd.concat(
                [base, pd.DataFrame(rows)], ignore_index=True
            )


# --- 1. НАСТРОЙКА СТРАНИЦЫ ---
st.set_page_config(page_title="Накладные ИП Саргсян", page_icon="📝", layout="wide")
st.title("📝 Система накладных")

# --- 2. ИНИЦИАЛИЗАЦИЯ ПАМЯТИ ---
if "main_data" not in st.session_state:
    st.session_state.main_data = _empty_positions_table()

top_form = st.container()
st.markdown("---")
table_area = st.container()
st.markdown("---")
actions_area = st.container()

# --- 3. ВЕРХНЯЯ ЧАСТЬ: ФОРМА ВВОДА (ДЛЯ ТЕЛЕФОНА) ---
with top_form:
    with st.expander("➕ ДОБАВИТЬ ПОЗИЦИЮ", expanded=True):
        with st.form("mobile_form", clear_on_submit=True):
            f_name = st.text_input("Наименование запчасти *")
            
            c1, c2, c3, c4 = st.columns([3, 2, 2, 1])
            with c1: f_art = st.text_input("Артикул")
            with c2: f_brand = st.text_input("Бренд")
            with c3: f_price = st.text_input("Цена (руб.)")
            with c4: f_qty = st.number_input("Кол-во", min_value=1, value=1)
            
            submit = st.form_submit_button("📥 Добавить в накладную", width="stretch")
            
            if submit:
                if not f_name.strip():
                    st.error("Пожалуйста, введите наименование!")
                else:
                    # Очистка цены от пробелов и запятых
                    try:
                        p_val = float(f_price.replace(',', '.').replace(' ', '').strip()) if f_price else 0.0
                    except (ValueError, TypeError):
                        p_val = 0.0
                    
                    new_row = pd.DataFrame([{
                        "Наименование": f_name.strip(),
                        "Артикул": f_art.strip(),
                        "Бренд": f_brand.strip(),
                        "Цена, руб.": p_val,
                        "Кол-во": f_qty
                    }])
                    st.session_state.main_data = pd.concat([st.session_state.main_data, new_row], ignore_index=True)
                    
                    # Сбрасываем кэш таблицы, чтобы она ровно перерисовалась
                    if _MAIN_EDITOR_KEY in st.session_state:
                        del st.session_state[_MAIN_EDITOR_KEY]
                    st.rerun()

# --- 4. НИЖНЯЯ ЧАСТЬ: РЕДАКТОР (ДЛЯ ПК / МАССОВОЙ ВСТАВКИ) ---
with table_area:
    st.markdown("### 📋 Список добавленных позиций")
    
    col_clear, _ = st.columns([1, 4])
    with col_clear:
        if not st.session_state.main_data.empty:
            if st.button("🗑️ Очистить таблицу", type="secondary", width="stretch"):
                st.session_state.main_data = _empty_positions_table()
                if _MAIN_EDITOR_KEY in st.session_state:
                    del st.session_state[_MAIN_EDITOR_KEY]
                st.rerun()

    # Правки в session_state подтягиваются в on_change до этого запуска скрипта
    # (см. _on_main_editor_change и issue Streamlit #7749).
    edited_df = st.data_editor(
        st.session_state.main_data,
        num_rows="dynamic",
        width="stretch",
        key=_MAIN_EDITOR_KEY,
        on_change=_on_main_editor_change,
    )

# --- 5. ИТОГИ И ГЕНЕРАЦИЯ EXCEL ---
with actions_area:
    def clean_num(val):
        try:
            return float(str(val).replace(',', '.').replace(' ', ''))
        except (ValueError, TypeError):
            return 0.0

    calc_df = edited_df.copy()
    calc_df['total'] = calc_df['Цена, руб.'].apply(clean_num) * calc_df['Кол-во'].apply(clean_num)
    total_sum = calc_df['total'].sum()
    pos_count = len(calc_df[calc_df["Наименование"].astype(str).str.strip() != ""])

    # Красивые счетчики
    m1, m2 = st.columns(2)
    m1.metric("Всего позиций", f"{pos_count} шт.")
    m2.metric("ИТОГО К ОПЛАТЕ", f"{total_sum:,.2f} руб.".replace(',', ' '))

    st.markdown("---")
    filename = st.text_input("Название файла (если нужно):", placeholder="Например: Иван_Бампер_Ауди")

    # Функция создания ИДЕАЛЬНОГО Excel
    def create_beautiful_excel(df):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Накладная"
        
        font_bold = Font(name='Arial', size=12, bold=True)
        font_reg = Font(name='Arial', size=12)
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        header_border = Border(left=Side(style='medium'), right=Side(style='medium'), top=Side(style='medium'), bottom=Side(style='medium'))
        center = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Шапка
        sheet['B2'] = "ИП Саргсян"; sheet['B2'].font = Font(name='Arial', size=14, bold=True)
        sheet['B4'] = "Расходная накладная"; sheet['B4'].font = Font(name='Arial', size=14, bold=True)
        sheet['B6'] = "Дата:"; sheet['B6'].font = font_bold
        sheet['C6'] = datetime.now().strftime('%d.%m.%Y'); sheet['C6'].font = font_bold

        # Заголовки
        headers = ["№", "Наименование", "Артикул", "Бренд", "Цена, руб.", "Кол-во", "Сумма, руб."]
        for i, h in enumerate(headers):
            c = sheet.cell(row=8, column=2+i, value=h)
            c.font = font_bold; c.border = header_border; c.alignment = center

        # Строки
        for i, row in df.iterrows():
            r_idx = 9 + i
            p = clean_num(row['Цена, руб.'])
            q = clean_num(row['Кол-во'])
            
            vals = [i+1, str(row['Наименование']), str(row['Артикул']), str(row['Бренд']), p, q, f"=F{r_idx}*G{r_idx}"]
            for j, v in enumerate(vals):
                cell = sheet.cell(row=r_idx, column=2+j, value=v)
                cell.font = font_reg; cell.border = border; cell.alignment = center
                if j in [4, 6]: cell.number_format = '#,##0.00' # Денежный формат
                    
        # Итого
        last_row = 9 + len(df)
        res_label = sheet.cell(row=last_row, column=7, value="Итого:")
        res_label.alignment = center; res_label.font = font_bold
        res_cell = sheet.cell(row=last_row, column=8, value=f"=SUM(H9:H{last_row-1})")
        res_cell.font = font_bold; res_cell.alignment = center; res_cell.border = border; res_cell.number_format = '#,##0.00'

        # Авто-ширина
        for col in range(2, 9):
            sheet.column_dimensions[get_column_letter(col)].width = 18

        buf = io.BytesIO()
        workbook.save(buf)
        buf.seek(0)
        return buf

    final_df = edited_df[edited_df["Наименование"].astype(str).str.strip() != ""]
    name_stripped = (filename or "").strip()
    fn = f"{name_stripped}.xlsx" if name_stripped else f"Накладная_{datetime.now().strftime('%H%M%S')}.xlsx"
    excel_payload = (
        create_beautiful_excel(final_df.reset_index(drop=True)).getvalue()
        if not final_df.empty
        else b""
    )
    st.download_button(
        "🚀 Скачать для Google Таблиц (.xlsx)",
        data=excel_payload,
        file_name=fn,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
        width="stretch",
        disabled=final_df.empty,
        help="Добавьте хотя бы одну строку с наименованием в таблице выше."
        if final_df.empty
        else None,
    )
